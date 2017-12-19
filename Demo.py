#!/usr/bin/env python
# coding: utf-8
#

from openpyxl import Workbook
from PyQt5.QtWidgets import (QWidget, QHBoxLayout,
                             QLabel, QApplication)
from PyQt5.QtGui import QPixmap


class Demo(QWidget):
    def __init__(self):
        # wb = Workbook()
        #
        # # grab the active worksheet
        # ws = wb.active
        #
        # # Data can be assigned directly to cells
        # ws['A1'] = 'ID'
        #
        # # Rows can also be appended
        # count = 1
        # for unit in range(101):
        #     count = count + 1
        #     ws['A%d' % count] = unit
        # # ws.append([1, 2, 3])
        #
        # # Save the file
        # wb.save("outputs.xlsx")
        #
        # print 'outputs.xlsx 导出成功!'

        QWidget.__init__(self)

        self.initUI()

    def initUI(self):
        hbox = QHBoxLayout(self)
        pixmap = QPixmap("img/1.png")

        lbl = QLabel(self)
        lbl.setPixmap(pixmap)

        hbox.addWidget(lbl)
        self.setLayout(hbox)

        self.move(300, 200)
        self.setWindowTitle('Red Rock')
        self.show()


def main():
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
