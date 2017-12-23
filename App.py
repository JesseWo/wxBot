#!/usr/bin/env python
# coding: utf-8
#
from datetime import date

from openpyxl import Workbook
import sys
from PyQt5.QtWidgets import (QWidget, QHBoxLayout,
                             QLabel, QApplication, QPushButton, QDesktopWidget, QBoxLayout, QVBoxLayout)
from PyQt5.QtGui import QPixmap, QIcon
from wxbot import *


class Application(WXBot):

    def on_qr_created(self, qr_file_path):
        self.init_ui(qr_file_path)

    def init_ui(self, qr_file_path):
        app = QApplication(sys.argv)

        widget = QWidget()
        hbox = QVBoxLayout(widget)
        # 提示文字
        warning_label = QLabel(widget)
        warning_label.setText('''声明:
        本软件仅用于个人测试使用, 严禁商业用途!
        请控制同一账号的使用频率, 过多使用可能会导致网页端限制登录
        Created by Jessewo''')
        hbox.addWidget(warning_label)
        # qr code
        self.image_label = QLabel(widget)
        pixmap = QPixmap(qr_file_path)
        self.image_label.setPixmap(pixmap)
        hbox.addWidget(self.image_label)

        self.tips_label = QLabel(widget)
        self.tips_label.setText('请扫描二维码登录微信')
        hbox.addWidget(self.tips_label)

        # 按钮
        self.btn_output = QPushButton('导出通讯录', widget)
        self.btn_output.clicked.connect(self.generate_xlsx)
        self.btn_output.setVisible(False)
        hbox.addWidget(self.btn_output)
        widget.setLayout(hbox)
        # 窗口居中显示
        qr = widget.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        widget.move(qr.topLeft())

        widget.setWindowTitle('Wechat ToolBox')
        widget.setWindowIcon(QIcon('img/icon.png'))
        widget.show()
        sys.exit(app.exec_())

    def on_login_success(self):
        self.tips_label.setText('登录成功!')
        self.image_label.setVisible(False)
        self.btn_output.setVisible(True)

    def generate_xlsx(self):
        # output all contacts to excel
        wb = Workbook()
        # grab the active worksheet
        ws = wb.active
        # Data can be assigned directly to cells
        ws['A1'] = 'index'
        ws['B1'] = 'UserName'
        ws['C1'] = 'NickName'
        ws['D1'] = 'PYInitial'
        ws['E1'] = 'PYQuanPin'
        ws['F1'] = 'Signature'
        ws['G1'] = 'RemarkName'
        ws['H1'] = 'Sex'
        ws['I1'] = 'Province'
        ws['J1'] = 'City'
        ws['K1'] = 'HeadImgUrl'
        ws['L1'] = 'StarFriend'
        ws['M1'] = 'KeyWord'
        ws['N1'] = 'AttrStatus'
        ws['O1'] = 'SnsFlag'
        ws['P1'] = 'MemberCount'
        ws['Q1'] = 'OwnerUin'
        ws['R1'] = 'ContactFlag'
        count = 1
        for contact in self.contact_list:
            count = count + 1
            # 序号
            ws['A%d' % count] = count - 1
            # username
            ws['B%d' % count] = contact['UserName']
            # 昵称
            ws['C%d' % count] = contact['NickName']
            # 昵称拼音简写
            ws['D%d' % count] = contact['PYInitial']
            # 昵称全拼
            ws['E%d' % count] = contact['PYQuanPin']
            # 签名
            ws['F%d' % count] = contact['Signature']
            # 备注名
            ws['G%d' % count] = contact['RemarkName']
            # 性别0默认 1男 2女
            ws['H%d' % count] = contact['Sex']
            # 地址
            ws['I%d' % count] = contact['Province']
            ws['J%d' % count] = contact['City']
            #
            ws['K%d' % count] = contact['HeadImgUrl']
            # other
            ws['L%d' % count] = contact['StarFriend']
            ws['M%d' % count] = contact['KeyWord']
            ws['N%d' % count] = contact['AttrStatus']
            ws['O%d' % count] = contact['SnsFlag']
            ws['P%d' % count] = contact['MemberCount']
            ws['Q%d' % count] = contact['OwnerUin']
            ws['R%d' % count] = contact['ContactFlag']
        # Save the file
        if not os.path.exists('outputs'):
            os.mkdir('outputs')
        wb.save('outputs/%s_%s.xlsx' % (self.my_account['NickName'], time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())))
        self.tips_label.setText('Excel导出成功!')
        print 'xlsx 导出成功!'


if __name__ == '__main__':
    bot = Application()
    bot.DEBUG = True
    bot.conf['qr'] = 'png'
    bot.run()
